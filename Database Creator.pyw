#!usr/bin/py
# -*- coding: utf-8 -*-
from Tkinter import *
import tkMessageBox
import openpyxl
import os
import multiprocessing
import tkFileDialog

#HERE WE DEFINE USEFUL AUXILIARY VARIABLES.
if __name__=="__main__":
 mess='-'*60+'\n'+' '*10+'Bienvenido al generador de base de datos 2017!\n'+\
 '-'*60+'\n'+' '*23+'INSTRUCCIONES DE USO\n'+\
 '\nPara comenzar, busque con el boton \'BUSCAR...\' la planilla con la cual va a a trabajar.\n'+\
 'Si no tiene una, el programa la creara por usted con el nombre que ingrese, en su Escritorio.\n'+\
 'Porfavor ingrese las fechas en el formato \'DD-MM-YYYY\'\n'+'-'*60+\
 '\nPara buscar un usuario en la planilla seleccionada ingrese su nombre y/o'+\
 ' apellido en el cuadro de abajo y se recibiran sus datos en la forma '+\
 '\'N# de Poliza, Rut, Nombre, Apellido '+'-'*60+'\n'+'\n'\
 +'Nombre base de la planilla a crear, si no tiene una (ej:\"Base de datos 1\"):'+'\n'*5\
 +'Nombre y/o apellido el cual buscar:'\
 +'\n'*20
 spacer=' '*7
 ls_frames=[]
 ls2=[['N# de Poliza:',spacer+'Rut:',spacer+'Nombre:',spacer+'Celular:'],['Tel.Oficina:',\
 spacer+'Tel:',spacer+'E-Mail:',spacer+'Dir:'],['Cumpleanos:',spacer+'Comision:',spacer+'Ejecutivo:',\
 spacer+'% Comision:'],['Compania:',spacer+'Inicio Vigencia:',spacer+'Fin Vigencia:']]
 ls_ents=[]
 ls_labs=[]
 workbook_list=[]
 User=[]
 number=0


#HERE WE INITIALIZE FRAMES AND THE MAIN WINDOW.
 def dir_ch():
    blkls=['Administrador','Publico','Public','defaultuser0','Usuario Publico','All Users','Default'
    ,'Default User']
    for i in os.listdir('C:\\Usuarios'):
        if i not in blkls:
            link='C:\\Usuarios\\'+str(i)
            User.append(link)
            return link

 root=Tk()
 root.title('Data Base Handler 2017')
 root.geometry('1550x650')
 root.wm_iconbitmap(dir_ch()+'\\Escritorio\\Database Creator\\Config\\favicon.ico')
 for i in range(4):
    ls_frames.append((Frame(root)))
 for i in ls_frames:
    i.grid(padx=100,pady=60)
 bottom_frame=Frame(root)
 bottom_frame.grid(row=4,column=1)
 right_frame=Frame(root)
 right_frame.grid(row=0,column=1,rowspan=15,columnspan=1)
 textbox= Text(right_frame,relief=SUNKEN,width='60',height='29',fg='red',wrap=WORD)
 textbox.insert(INSERT,mess)
 textbox.pack()
 textbox.mark_set('input1_start','12.75')#12.71
 textbox.mark_set('input1_end','16.99')#17.0
 textbox.mark_set('input2_start','17.35')
 textbox.mark_set('input2_end','27.0')


#HERE GOES FUNCTIONS AND ACTIONS.

def packer_dest1():
      M3.pack(side=BOTTOM)
      M2.pack_forget()
      return 0

def packer_dest2():
      M2.pack(side=BOTTOM)
      M3.pack_forget()
      return 0


def X_Proc(dr,fil,send_end):
    flag=False
    black=['My Music','My Pictures','My Videos']
    ls=fsanitizer(dr,black)
    for i in ls:
        path=dr
        if i==fil:
            flag=True
            send_end.send((flag,path+'\\'+fil))

            return 0
        path=path+'\\'+i
        #if os.path.isdir(path)==True:
         #  X_Proc(path,fil,send_end)
    return None

def fsanitizer(fold,bl):
    san=[]
    for f in os.listdir(fold):
        if os.path.isdir(fold+'\\'+f)==True and f not in bl:
             san.append(f)
    return san

def file_searcher(fil):

    blk=['AppData','Application Data','Cookies','Local Settings','NetHood','PrintHood'
    ,'SendTo','Start Menu','Templates','My Documents','Recent']
    dr=fsanitizer(User[0],blk)
    print dr
    ls_proc=[]
    pipe_list=[]
    for n in range(len(dr)):
          recv_end,send_end = multiprocessing.Pipe()
          ls_proc.append(multiprocessing.Process(target=X_Proc,args=(User[0]+'\\'+dr[n],fil,send_end)))
          pipe_list.append(recv_end)
          ls_proc[n].start()
    print 'its the processes'
    for proc in ls_proc:
        proc.join()
        print 'yep, its the processes: '+str(proc)
    print "nigga?!"
    print pipe_list
    result_list=[]
    for pipe in pipe_list:
        result_list.append(pipe.recv())
    #result_list=[x.recv() for x in pipe_list]
    print result_list
    return result_list

def aux_ls(ls):
    nls=[None,'TIPO DE SEGURO']
    aux=[]
    for i in ls:
        aux+=i
    for i in aux:
        nls.append(i.strip(": ").upper())
    return nls

def file_creater(inp):
    ln=User[0]+'\\Escritorio\\'+inp+'.xlsx'
    wb=openpyxl.Workbook()
    sheet=wb.active
    sheet.freeze_panes='A2'
    new_ls=aux_ls(ls2)
    fn=openpyxl.styles.Font(bold=True)
    for i in range(1,17):
      sheet.cell(row=1,column=i).font=fn
      sheet.cell(row=1,column=i,value=new_ls[i])
    for i in 'ABCDEFGHIJKLMNOP':
        sheet.column_dimensions[i].width=20
    wb.save(ln)
    #print 'new file created at desktop!'
    return ln

#TENGO QUE CAMBIAR ESTA FUNCION RESPECTO A LA NUEVA, MULTIPROCESSING BIEN IMPLEMENTADO
#(REVISAR EN TO DO CASO)

def is_inthere(inp):
     flag=False
     r_ls=file_searcher(inp+'.xlsx')
     print r_ls
     for i,y in r_ls:
         if i==True:
            flag,path=i,y
            break
     if flag==True:
        workbook_list=[]
        workbook_list.append(path)
        #print 'workbook selected!'
     else:
         file_creater(inp)
         cleaner(0,1)
         #print 'ready to work at new workbook'
     return flag

def submit():
     global seg
     if len(workbook_list)==0 or workbook_list[0]=='':
        tkMessageBox.showerror('ERROR','No se ha seleccionado una planilla en la cual trabajar.')
        return 0
     if seg=='':
         tkMessageBox.showerror('ERROR','No se ha seleccionado el tipo de seguro.')
         return 0
     path=workbook_list[0]
     wb=openpyxl.load_workbook(path)
     sheet=wb.active
     sheet.title='Clientes'
     start=sheet.max_row+1
     ls_text=[None,seg]
     for i in range(15):
         if ls_ents[i].get()=='':
            ls_text.append('No Hay Datos')
         else:
            ls_text.append(ls_ents[i].get())
     for i in range(1,17):
         sheet.cell(row=start,column=i,value=ls_text[i])
     for i in range(15):
         ls_ents[i].delete(0,last=99)
     #if radlist==2:
    #    for i in range(9):
    #        rad2[i].deselect()
     #else:
    #    for i in range(5):
    #        rad3[i].deselect()
     seg=''
     try:
         wb.save(path)
     except IOError:
         tkMessageBox.showerror("ERROR","Los datos no han podido ser guardados,"+
         " por favor cierre el archivo antes de modificarlo.")
         return 0
     return 0

def name_aux(word):
    words=word.split(' ')
    nwords=[]
    flag,flag2=False,False
    for i in words:
        if i!=u'':
            nwords.append(i)
    if len(nwords)==4:
        nwords=[nwords[0],nwords[2],nwords[3]]
    elif len(nwords)==2:
          flag=True
          if nwords[0][0].isupper() or nwords[1][0].isupper():
              nwords[0],nwords[1]=nwords[0].lower(),nwords[1].lower()
              flag2=True
    if len(nwords)==3:
          flag=True
          if nwords[0][0].isupper() or nwords[1][0].isupper() or nwords[2][0].isupper():
              nwords[0],nwords[1],nwords[2]=nwords[0].lower(),nwords[1].lower(),nwords[2].lower()
              flag2=True
    if flag==True:
      word2=' '.join(nwords)
      if flag2!=True:
         wd=[]
         for string in nwords:
             start=string[0].upper()
             end=string[1:]
             wd.append(start+end)
         word2=' '.join(wd)

    elif len(nwords)==1 and nwords[0]!=u'':
       if nwords[0][0].isupper():
           word2=nwords[0].lower()
       else:
           word2=nwords[0][0].upper()+nwords[0][1:]
    else:
       return 1
    return [word,word2]

def next_name(string):
    strings=string.split('\n')
    nstr=strings[0][8:]
    return nstr

def name_aux2(name):
    names=name.split(' ')
    if len(names)==3:
        names=[names[0],names[1]]
    if len(names)==2:
       return [names[0],names[1]]
    else:
       return [names[0],False]

def name_looker():
    global number
    flag,flag2,flag3=False,False,True
    if len(workbook_list)==0:
        tkMessageBox.showerror('ERROR','no ha seleccionado una planilla en la cual buscar')
        return 0
    name=textbox.get('17.35','27.0').strip('\n')
    if u'Nombre: ' in name:
        flag2=True
        name=next_name(name)
    strings=name_aux(name)
    if flag2==True:
        fname1,sname1=name_aux2(strings[0])
        fname2,sname2=name_aux2(strings[1])
        if sname1==False:
            flag3=False
    if strings==1:
        tkMessageBox.showerror('ERROR','Input invalido, por favor ingrese el nombre con mas/menos palabras.')
        return 0
    wb=openpyxl.load_workbook(workbook_list[0])
    sheet=wb.active
    shlong=sheet.max_row+1
    for i in range(1,shlong):
        expr=sheet.cell(row=i,column=4).value
        if flag2!=True:
          if strings[0] in expr or strings[1] in expr:
            flag=True
            break
        else:
            if (sname1 in expr and expr!=name and i>number) or (sname2 in expr and expr!=name and i>number):
                 flag=True
                 break
    if flag==True:
        number=i
        cleaner(2,3)
        textbox.insert('19.0','Nombre: '+expr)
        textbox.insert('20.0','Celular: '+sheet.cell(row=i,column=5).value)
        textbox.insert('21.0','E-Mail: '+sheet.cell(row=i,column=8).value)
        textbox.insert('22.0','Tipo de Seguro: '+sheet.cell(row=i,column=1).value)
        textbox.insert('23.0','N# Poliza: '+sheet.cell(row=i,column=2).value)
        textbox.insert('24.0','RUT: '+sheet.cell(row=i,column=3).value)
        textbox.insert('25.0','Cumpleanos: '+sheet.cell(row=i,column=10).value)
        textbox.insert('26.0','Fin Vigencia: '+sheet.cell(row=i,column=16).value)
    else:
        cleaner(2,3)
        tkMessageBox.showerror('Intente Otra Vez','No se ha encontrado a la persona, pruebe con mayor'+
        ' precision poniendo mas apellidos o verifique que el nombre que ha puesto este bien escrito.')
    return 0

def rad_get():
    global seg
    seg=str(var.get())
    global radlist
    if seg in ls_m2:
        radlist=2
    else:
        radlist=3
    return 0

def path_sanitizer(cpat):
    cpat=str(cpat)
    cpat=cpat.split('/')
    cpat='\\'.join(cpat)
    return cpat

def sheet_searcher():
    inp=textbox.get('12.75','17.0')
    User=[]
    dir_ch()
    string=str(inp.strip())
    if len(string)==0 or string=='ready to work at new workbook' or string=='workbook selected!':
       path=tkFileDialog.askopenfilename(filetypes = (("Template files", "*.xlsx"), ("All files", "*")))
       path=path_sanitizer(path)
       global workbook_list
       workbook_list=[]
       workbook_list.append(path)
       cleaner(0,1)
       if path!='':
         textbox.insert('14.0','workbook selected!')

    else:
      file_creater(string)
      cleaner(0,1)
      textbox.insert('14.0','ready to work at new workbook')

    return None

def cleaner(x,y):
    ls=['12.75','17.0','17.35','27.0']

    textbox.delete(ls[x],ls[y])
    if x!=2:
      textbox.insert(ls[x],'\n'*5)
    else:
      textbox.insert(ls[x],'\n'*11)
    return 0



def tester():
    string=textbox.get('input1_start','input1_end')
    for i in string:
       if i !=' ' and i!='\n':
          print('nigga!')
    textbox.insert('12.75','nigga')
    print len(string)
    return 0

def clean():
    textbox.delete('12.75','17.0')
    textbox.insert('12.75','\n'*5)
    textbox.delete('17.35','27.0')
    textbox.insert('17.35','\n'*11)
    return 0

#HERE WE INITIALIZE OBJECTS.
if __name__=="__main__":
 for e in range(3):
   for i in range(4):
     ls_labs.append((Label(ls_frames[e],text=ls2[e][i])))
     ls_ents.append((Entry(ls_frames[e])))
 for i in range(3):
     ls_labs.append((Label(ls_frames[3],text=ls2[3][i])))
     ls_ents.append((Entry(ls_frames[3])))

 ls_m2=['Vehiculos Livianos','Vehiculos Pesados','Incendio y Sismo','Accidentes Personales',
 'Riesgos Comerciales','Responsabilidad Civil','Equipos Moviles',
 'Riesgo en Construccion','Asistencia de Viaje']
 ls_m3=['APV','Seguros Flexibles','Full Ahorro','Patrimonial','Seguro Salud']
 seg=''
 rad2=[]
 rad3=[]
 var=StringVar()

 M2=Menubutton(ls_frames[3],text='Seguros Generales',activeforeground="white",relief=SUNKEN,height=1)
 M2.menu=Menu(M2,tearoff=0)
 M2['menu']=M2.menu
 for i in range(9):
     rad2.append(M2.menu.add_radiobutton(label=ls_m2[i],variable=var,value=ls_m2[i],command=rad_get))

 M3=Menubutton(ls_frames[3],text='Seguros de Vida',activeforeground="white",relief=SUNKEN,height=1)
 M3.menu=Menu(M3,tearoff=0)
 M3['menu']=M3.menu
 for i in range(5):
     rad3.append(M3.menu.add_radiobutton(label=ls_m3[i],variable=var,value=ls_m3[i],command=rad_get))

 M1=Menubutton(ls_frames[3],text='Tipo de Seguro',relief=SUNKEN)
 M1.grid(sticky=NE)
 M1.menu=Menu(M1,tearoff=0)
 M1['menu']=M1.menu
 M1.menu.add_radiobutton(label='Seguros de Vida',command=packer_dest1)
 M1.menu.add_radiobutton(label='Seguros Generales',command=packer_dest2)

 SHEET=Button(bottom_frame,text='BUSCAR PLANILLA',relief=RAISED,command=sheet_searcher)
 SEARCH=Button(bottom_frame,text='BUSCAR',relief=RAISED,command=name_looker)
 SUBMIT= Button(bottom_frame, text="SUBMIT", relief=RAISED, activebackground="red",command=submit)
 CLEANSE=Button(bottom_frame,text='CLEAN',relief=RAISED,command=clean)

#HERE WE PACK,PLACE AND GRID THOSE OBJECTS.

 for i in range(15):
     ls_labs[i].pack(side=LEFT)
     ls_ents[i].pack(side=LEFT)
#ENT.pack(side=RIGHT)
#L1.pack(side=RIGHT)
 SHEET.pack(side=LEFT)
 SUBMIT.pack(side=RIGHT)
 CLEANSE.pack(side=RIGHT)
 SEARCH.pack(side=LEFT)
 M1.pack()
 M1.place()
#B1.pack(side=LEFT)
#B2.pack(side=RIGHT)
 root.mainloop()
